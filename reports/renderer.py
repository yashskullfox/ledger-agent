"""
reports/renderer.py  –  Console + file rendering for all financial reports
───────────────────────────────────────────────────────────────────────────
Uses `rich` for beautiful console output and pandas + openpyxl for Excel export.
"""
from __future__ import annotations

import csv
import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional

from accounting.balance_sheet import BalanceSheet
from config import EXPORTS_DIR, REPORT_DATE_FMT
from core.models import COAType, Transaction

try:
    from rich import box
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text

    _RICH = True
except ImportError:
    _RICH = False

try:
    import pandas as pd

    _PANDAS = True
except ImportError:
    _PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

console = Console() if _RICH else None

# Colour palette
C_HEADER = "bold cyan"
C_SUBTOTAL = "bold yellow"
C_POSITIVE = "green"
C_NEGATIVE = "red"
C_SECTION = "bold white on dark_blue"
C_BALANCED = "bold green"
C_UNBAL = "bold red"


def _fmt(amount: Decimal, width: int = 14) -> str:
    if amount < 0:
        return f"({abs(amount):>{width - 2},.2f})"
    return f"{amount:>{width},.2f}"


def _fmt_rich(amount: Decimal) -> Text:
    if not _RICH:
        return _fmt(amount)  # type: ignore
    s = _fmt(amount)
    colour = C_NEGATIVE if amount < 0 else C_POSITIVE
    return Text(s, style=colour)


def render_balance_sheet(bs: BalanceSheet, show_income: bool = True) -> None:
    if _RICH:
        _render_bs_rich(bs, show_income)
    else:
        _render_bs_plain(bs, show_income)


def _render_bs_rich(bs: BalanceSheet, show_income: bool) -> None:
    now = datetime.now().strftime(REPORT_DATE_FMT)
    console.print()
    console.rule(f"[{C_HEADER}] {bs.entity_name}  ·  Balance Sheet  ·  {bs.period} [{C_HEADER}]")
    console.print(f"  [dim]Generated: {now}[/dim]\n")

    tbl = Table(show_header=False, box=box.SIMPLE_HEAD, padding=(0, 1),
                show_edge=False, width=80)
    tbl.add_column("Label", style="white", no_wrap=False, ratio=65)
    tbl.add_column("Amount", style="white", justify="right", ratio=35)

    sections = [
        ("ASSETS", COAType.ASSET, C_SECTION),
        ("LIABILITIES", COAType.LIABILITY, C_SECTION),
    ]
    if show_income:
        sections += [
            ("REVENUE", COAType.REVENUE, C_SECTION),
            ("EXPENSES", COAType.EXPENSE, C_SECTION),
        ]
    sections += [("MEMBERS' EQUITY", COAType.EQUITY, C_SECTION)]

    for section_label, coa_type, sec_style in sections:
        tbl.add_row(Text(f" {section_label}", style=sec_style), "")
        lines = [l for l in bs.lines if l.coa_type == coa_type]
        for line in lines:
            indent = "  " * line.indent
            label = f"{indent}{line.label}"
            if line.is_subtotal:
                tbl.add_row(
                    Text(label, style=C_SUBTOTAL),
                    Text(_fmt(line.amount), style=C_SUBTOTAL),
                )
                tbl.add_row("", "")
            elif line.amount == 0 and not line.is_subtotal:
                tbl.add_row(Text(label, style="bold"), "")
            else:
                tbl.add_row(label, _fmt_rich(line.amount))
        tbl.add_row("", "")

    # Net income line
    if show_income:
        tbl.add_row(Text("  Net Income (Revenue − Expenses)", style="bold"),
                    _fmt_rich(bs.net_income))
        tbl.add_row("", "")

    console.print(tbl)

    # Balance check
    diff = abs(bs.total_assets - (bs.total_liabilities + bs.total_equity))
    if bs.is_balanced:
        console.print(f"  [{C_BALANCED}]✓ BALANCED[/{C_BALANCED}]  "
                      f"Assets ${bs.total_assets:,.2f}  =  "
                      f"Liabilities ${bs.total_liabilities:,.2f}  +  "
                      f"Equity ${bs.total_equity:,.2f}")
    else:
        console.print(f"  [{C_UNBAL}]⚠ OUT OF BALANCE  Difference: ${diff:,.2f}[/{C_UNBAL}]")
    console.print()


def _render_bs_plain(bs: BalanceSheet, show_income: bool) -> None:
    w = 72
    print("=" * w)
    print(f"  {bs.entity_name}  —  Balance Sheet  —  {bs.period}")
    print("=" * w)
    sections = [
        ("ASSETS", [l for l in bs.lines if l.coa_type == COAType.ASSET]),
        ("LIABILITIES", [l for l in bs.lines if l.coa_type == COAType.LIABILITY]),
    ]
    if show_income:
        sections += [
            ("REVENUE", [l for l in bs.lines if l.coa_type == COAType.REVENUE]),
            ("EXPENSES", [l for l in bs.lines if l.coa_type == COAType.EXPENSE]),
        ]
    sections += [("MEMBERS' EQUITY", [l for l in bs.lines if l.coa_type == COAType.EQUITY])]

    for sec_label, lines in sections:
        print(f"\n  {sec_label}")
        print("  " + "-" * (w - 2))
        for line in lines:
            indent = "  " * line.indent
            label = f"{indent}{line.label}"
            amt = _fmt(line.amount)
            sep = "─" * 3 if line.is_subtotal else " "
            print(f"  {label:<50}  {amt:>14}")
    print("\n" + "=" * w)
    bal_label = "✓ BALANCED" if bs.is_balanced else "⚠ OUT OF BALANCE"
    print(f"  {bal_label}   Assets: ${bs.total_assets:,.2f}  =  "
          f"Liabilities: ${bs.total_liabilities:,.2f}  +  "
          f"Equity: ${bs.total_equity:,.2f}")
    print("=" * w)


def render_transactions(
        transactions: List[Transaction],
        title: str = "Transactions",
        show_coa: bool = True,
) -> None:
    if not _RICH:
        _render_txn_plain(transactions, title, show_coa)
        return

    tbl = Table(title=title, box=box.ROUNDED, show_lines=False)
    tbl.add_column("Date", style="dim", width=12)
    tbl.add_column("Description", style="white", max_width=40)
    if show_coa:
        tbl.add_column("COA", style="cyan", width=10)
    tbl.add_column("Amount", justify="right", width=14)
    tbl.add_column("Type", style="dim", width=18)

    for t in transactions:
        amt_text = _fmt_rich(t.amount)
        row = [str(t.date), t.description[:40]]
        if show_coa:
            row.append(t.coa_code or "?")
        row += [amt_text, t.transaction_type.value]
        tbl.add_row(*row)

    console.print(tbl)


def _render_txn_plain(txns, title, show_coa):
    print(f"\n{title}")
    print("-" * 80)
    for t in txns:
        coa = f"[{t.coa_code}]" if show_coa else ""
        print(f"  {t.date}  {_fmt(t.amount)}  {coa:<10}  {t.description[:50]}")
    print("-" * 80)


def export_balance_sheet_csv(bs: BalanceSheet, out_dir: Path = EXPORTS_DIR) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = out_dir / f"balance_sheet_{bs.period}.csv"
    with open(fname, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Entity", bs.entity_name])
        writer.writerow(["EntityID", bs.entity_id])  # ARCH-30: scope artefact to entity
        writer.writerow(["Period", bs.period])
        writer.writerow(["Generated", datetime.now().isoformat()])
        writer.writerow([])
        writer.writerow(["Code", "Label", "Type", "Amount", "Is Subtotal"])
        for line in bs.lines:
            writer.writerow([
                line.coa_code, line.label,
                line.coa_type.value,
                str(line.amount),
                "Y" if line.is_subtotal else "",
            ])
        writer.writerow([])
        writer.writerow(["Total Assets", str(bs.total_assets)])
        writer.writerow(["Total Liabilities", str(bs.total_liabilities)])
        writer.writerow(["Total Equity", str(bs.total_equity)])
        writer.writerow(["Net Income", str(bs.net_income)])
        writer.writerow(["Balanced?", "YES" if bs.is_balanced else "NO"])
    return fname


def export_transactions_csv(
        transactions: List[Transaction],
        period: str,
        out_dir: Path = EXPORTS_DIR,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = out_dir / f"transactions_{period}.csv"
    with open(fname, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Date", "Description", "Amount", "Type",
            "COA Code", "COA Name", "Is Transfer", "Tags", "Notes",
        ])
        for t in transactions:
            writer.writerow([
                t.date, t.description, str(t.amount),
                t.transaction_type.value,
                t.coa_code, t.coa_name,
                "Y" if t.is_transfer else "N",
                ", ".join(t.tags),
                t.notes,
            ])
    return fname


def export_balance_sheet_excel(bs: BalanceSheet,
                               out_dir: Path = EXPORTS_DIR) -> Optional[Path]:
    if not _OPENPYXL:
        print("openpyxl not installed – skipping Excel export")
        return None

    out_dir.mkdir(parents=True, exist_ok=True)
    fname = out_dir / f"balance_sheet_{bs.period}.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = f"Balance Sheet {bs.period}"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    subtotal_font = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="2E75B6")
    section_font = Font(color="FFFFFF", bold=True)

    def _add_header(text: str, col_span: int = 3):
        row = ws.max_row + 1
        ws.cell(row, 1, text).font = section_font
        ws.cell(row, 1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)

    # Title
    ws.cell(1, 1, f"{bs.entity_name}  —  Balance Sheet  —  {bs.period}").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Generated: {datetime.now().strftime(REPORT_DATE_FMT)}").font = Font(italic=True)
    ws.append([])
    ws.append(["Code", "Description", "Amount"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    sections = [
        ("ASSETS", COAType.ASSET),
        ("LIABILITIES", COAType.LIABILITY),
        ("REVENUE", COAType.REVENUE),
        ("EXPENSES", COAType.EXPENSE),
        ("MEMBERS' EQUITY", COAType.EQUITY),
    ]
    for sec_label, coa_type in sections:
        _add_header(sec_label)
        for line in (l for l in bs.lines if l.coa_type == coa_type):
            indent = "  " * line.indent
            row_n = ws.max_row + 1
            ws.cell(row_n, 1, line.coa_code)
            ws.cell(row_n, 2, f"{indent}{line.label}")
            ws.cell(row_n, 3, float(line.amount))
            ws.cell(row_n, 3).number_format = '#,##0.00'
            ws.cell(row_n, 3).alignment = Alignment(horizontal="right")
            if line.is_subtotal:
                for c in range(1, 4):
                    ws.cell(row_n, c).font = subtotal_font

    # Summary
    ws.append([])
    ws.append(["", "TOTAL ASSETS", float(bs.total_assets)])
    ws.append(["", "TOTAL LIABILITIES", float(bs.total_liabilities)])
    ws.append(["", "TOTAL EQUITY", float(bs.total_equity)])
    ws.append(["", "NET INCOME", float(bs.net_income)])
    ws.append(["", "BALANCED?", "YES" if bs.is_balanced else "NO"])

    wb.save(str(fname))
    return fname


def export_balance_sheet_json(bs: BalanceSheet,
                              out_dir: Path = EXPORTS_DIR) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = out_dir / f"balance_sheet_{bs.period}.json"
    payload = {
        "entity_id": bs.entity_id,  # ARCH-30: scope artefact to entity
        "entity_name": bs.entity_name,
        "period": bs.period,
        "generated_at": datetime.now().isoformat(),
        "total_assets": str(bs.total_assets),
        "total_liabilities": str(bs.total_liabilities),
        "total_equity": str(bs.total_equity),
        "net_income": str(bs.net_income),
        "is_balanced": bs.is_balanced,
        "lines": [
            {
                "code": l.coa_code,
                "label": l.label,
                "type": l.coa_type.value,
                "amount": str(l.amount),
                "is_subtotal": l.is_subtotal,
                "indent": l.indent,
            }
            for l in bs.lines
        ],
    }
    fname.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    # ARCH-26: emit coverage manifest alongside every balance sheet export
    _emit_coverage_manifest(bs, out_dir)
    return fname


def _emit_coverage_manifest(bs: BalanceSheet, out_dir: Path) -> Path:
    """ARCH-26: Write a sibling *.coverage.json listing consumed/skipped snapshots.

    The manifest is consumed by the 12-month gate (R-45) to detect silent drops.
    Any account that has a snapshot for the period is "consumed"; accounts without
    a snapshot for this period are "skipped" (AggregationGap warning).
    """
    try:
        from core.database import AccountRepo, SnapshotRepo
        accounts = AccountRepo.list_for_entity(bs.entity_id) if bs.entity_id else []
        snapshots_for_period = {
            s.account_id
            for s in SnapshotRepo.list_for_entity(bs.entity_id)
            if s.statement_period == bs.period
        } if bs.entity_id else set()

        consumed = []
        skipped = []
        for acct in accounts:
            entry = {"account_id": acct.id, "institution": acct.institution,
                     "name": acct.name, "period": bs.period}
            if acct.id in snapshots_for_period:
                consumed.append(entry)
            else:
                skipped.append({**entry, "reason": "AggregationGap: no snapshot for period"})

        manifest = {
            "entity_id": bs.entity_id,
            "entity_name": bs.entity_name,
            "period": bs.period,
            "generated_at": datetime.now().isoformat(),
            "consumed_snapshots": consumed,
            "skipped_snapshots": skipped,
            "gap_count": len(skipped),
        }
    except Exception as exc:
        manifest = {
            "entity_id": bs.entity_id,
            "entity_name": bs.entity_name,
            "period": bs.period,
            "generated_at": datetime.now().isoformat(),
            "error": str(exc),
            "consumed_snapshots": [],
            "skipped_snapshots": [],
            "gap_count": 0,
        }

    cov_fname = out_dir / f"balance_sheet_{bs.period}.coverage.json"
    cov_fname.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return cov_fname


# ── Tax estimate renderer (moved here from accounting.tax_estimator – ARCH-02) ──

def render_tax_estimate(est) -> None:  # type: ignore[override]
    """Print a Rich-formatted tax estimate table to stdout.

    Accepts a ``TaxEstimate`` object from ``accounting.tax_estimator``.
    Imported lazily so callers without Rich still get plain-text output.
    """
    try:
        from rich.console import Console
        from rich.panel import Panel
        from rich.table import Table

        console = Console()
        console.print(Panel(
            f"[bold yellow]Tax Estimate – {est.entity_name}[/bold yellow]\n"
            f"[dim]Period: {est.period}  |  "
            f"Effective rate: {est.effective_rate}%[/dim]",
            border_style="yellow",
        ))

        tbl = Table(show_header=True, header_style="bold cyan")
        tbl.add_column("Component", style="white", min_width=30)
        tbl.add_column("Annual Amount", style="yellow", justify="right", min_width=16)
        tbl.add_column("Quarterly Payment", style="green", justify="right", min_width=16)

        for label, annual, quarterly in [
            ("Net Income (annualized)", est.net_income, est.net_income / 4),
            ("Self-Employment Tax (15.3%)", est.se_tax, est.se_tax / 4),
            ("Federal Income Tax (~22%)", est.federal_income_tax, est.federal_income_tax / 4),
            ("State Income Tax", est.state_income_tax, est.state_income_tax / 4),
        ]:
            tbl.add_row(label, f"${annual:,.2f}", f"${quarterly:,.2f}")

        tbl.add_section()
        tbl.add_row(
            "[bold]TOTAL ESTIMATED TAX[/bold]",
            f"[bold]${est.total_annual_tax:,.2f}[/bold]",
            f"[bold]${est.total_annual_tax / 4:,.2f}[/bold]",
        )
        console.print(tbl)
        console.print("\n[bold cyan]Quarterly Payment Schedule[/bold cyan]")
        for pmt in est.quarterly_payments:
            console.print(
                f"  [yellow]{pmt.quarter}[/yellow]  "
                f"Due: [dim]{pmt.due_date}[/dim]  →  "
                f"[green]${pmt.amount:,.2f}[/green]"
            )
        console.print()
        for note in est.notes:
            console.print(f"  [dim]{note}[/dim]")

    except ImportError:
        print(f"\n=== Tax Estimate – {est.entity_name} ({est.period}) ===")
        print(f"Net Income (annual): ${est.net_income:,.2f}")
        print(f"Total Tax:           ${est.total_annual_tax:,.2f}")
        print(f"Effective Rate:      {est.effective_rate}%")
        for pmt in est.quarterly_payments:
            print(f"  {pmt.quarter}  {pmt.due_date}: ${pmt.amount:,.2f}")
        for note in est.notes:
            print(f"  {note}")
